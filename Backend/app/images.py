from flask import Blueprint, request, send_file, jsonify
import openpyxl
from io import BytesIO

images_bp = Blueprint('images', __name__)

# Load Excel once on startup
excel_file = "zla_chatbot.xlsx"
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Helper: Extract headers
headers = {cell.value.strip().lower(): idx for idx, cell in enumerate(ws[1])}

@images_bp.route("/car-image", methods=["GET"])
def car_image():
    make = request.args.get("make")
    model = request.args.get("model")
    year = request.args.get("year")

    if not all([make, model, year]):
        return jsonify({"error": "Missing 'make', 'model' or 'year' parameter"}), 400

    # Find matching row
    matched_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        mk = str(row[headers['make']].value).strip().lower()
        mod = str(row[headers['model']].value).strip().lower()
        yr = str(row[headers['year']].value).strip()

        if mk == make.lower() and mod == model.lower() and yr == str(year):
            matched_row = row_idx
            break

    if matched_row is None:
        return jsonify({"error": "No matching row found for the provided values."}), 404

    # Find image in that row
    for image in ws._images:
        anchor_row = image.anchor._from.row + 1
        if anchor_row == matched_row:
            image_stream = BytesIO(image._data())
            image_stream.seek(0)
            return send_file(image_stream, mimetype='image/png')

    return jsonify({"error": "Match found but no image attached to that row."}), 404
